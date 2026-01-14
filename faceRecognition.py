import os
import time
from datetime import datetime

import cv2
import numpy as np
from openpyxl import Workbook, load_workbook

haar_file = 'haarcascade_frontalface_default.xml'
face_cascade = cv2.CascadeClassifier(haar_file)
datasets = 'dataset'
width, height = 130, 100
log_file = 'recognition_log.xlsx'

#
def append_log(name: str, confidence: float | None):
    if not os.path.exists(log_file):
        wb = Workbook()
        ws = wb.active
        ws.title = 'log'
        ws.append(['timestamp', 'name', 'confidence'])
        wb.save(log_file)

    wb = load_workbook(log_file)
    ws = wb.active
    ws.append([datetime.now().isoformat(timespec='seconds'), name, confidence if confidence is not None else ''])
    wb.save(log_file)

if not hasattr(cv2, "face"):
    raise SystemExit("OpenCV was built without the face module. Install opencv-contrib-python and try again.")

print('Training...')
images, labels, names = [], [], {}
last_log_time = {}
logged_once = False

for person_id, subdir in enumerate(sorted(os.listdir(datasets))):
    subjectpath = os.path.join(datasets, subdir)
    if not os.path.isdir(subjectpath):
        continue
    names[person_id] = subdir
    for filename in os.listdir(subjectpath):
        path = os.path.join(subjectpath, filename)
        img = cv2.imread(path, cv2.IMREAD_GRAYSCALE)
        if img is None:
            continue
        img = cv2.resize(img, (width, height))
        images.append(img)
        labels.append(int(person_id))

if not images:
    raise SystemExit("No training images found in dataset. Run main.py to create some first.")

images = np.array(images)
labels = np.array(labels)
model = cv2.face.LBPHFaceRecognizer_create()
#model =  cv2.face.FisherFaceRecognizer_create()

model.train(images, labels)

webcam = cv2.VideoCapture(0)
cnt=0

while True:
    ret, im = webcam.read()
    if not ret or im is None:
        print("Failed to capture frame; exiting.")
        break
    gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, 1.3, 5)
    for (x,y,w,h) in faces:
        cv2.rectangle(im,(x,y),(x+w,y+h),(255,255,0),2)
        face = gray[y:y + h, x:x + w]
        face_resize = cv2.resize(face, (width, height))

        prediction = model.predict(face_resize)
        conf = float(prediction[1])
        cv2.rectangle(im, (x, y), (x + w, y + h), (0, 255, 0), 3)
        if conf < 800:
            name = names[prediction[0]]
            cv2.putText(im, '%s - %.0f' % (name, conf), (x-10, y-10), cv2.FONT_HERSHEY_PLAIN, 2, (0, 0, 255))
            print(name)
            if not logged_once:
                append_log(name, conf)
                logged_once = True
            cnt=0
        
        else:
            cnt+=1
            cv2.putText(im,'Unknown',(x-10, y-10), cv2.FONT_HERSHEY_PLAIN,1,(0, 255, 0))
            if(cnt>100):
                print("Unknown Person")
                if not logged_once:
                    append_log('Unknown', None)
                    logged_once = True
                cv2.imwrite("unKnown.jpg",im)
                cnt=0
    cv2.putText(im, 'Press q or ESC to exit', (10, 25), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 255, 255), 2)
    cv2.imshow('FaceRecognition', im)
    key = cv2.waitKey(10) & 0xFF
    if key in (27, ord('q')):
        break

webcam.release()
cv2.destroyAllWindows()



